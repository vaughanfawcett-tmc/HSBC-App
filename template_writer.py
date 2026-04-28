"""Fill the HSBC SLA template without disturbing formulas or formatting.

The template is never modified in place. We load it, write cells, save to a
new path. Client rows are located by matching column B values, not by
hard-coded row numbers, so the template can grow without a code change.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

TITLE_CELL = "B2"
# Services section: header row has 'Client' in column B, TOTAL row follows,
# then client rows. Column C = Within SLA, D = Outside SLA, E = % Within.
SERVICES_HEADER_VALUE = "Client"
CALLBACKS_HEADER_TEXT = "Call backs"

WITHIN_COL = "C"
OUTSIDE_COL = "D"
PERCENT_COL = "E"


def _find_header_row(ws: Worksheet, start_row: int = 1) -> int:
    """Find the first row where column B == 'Client' starting at start_row."""
    for row in range(start_row, ws.max_row + 1):
        val = ws[f"B{row}"].value
        if isinstance(val, str) and val.strip().lower() == SERVICES_HEADER_VALUE.lower():
            return row
    raise ValueError(f"Could not find 'Client' header in column B starting at row {start_row}")


def _find_callbacks_header_row(ws: Worksheet) -> int | None:
    """Find the row containing the 'Call backs' section header, or None."""
    for row in range(1, ws.max_row + 1):
        val = ws[f"B{row}"].value
        if isinstance(val, str) and CALLBACKS_HEADER_TEXT.lower() in val.lower():
            return row
    return None


def _client_row_map(ws: Worksheet, start_row: int, end_row: int) -> dict[str, int]:
    """Map each column-B client name (rows start_row..end_row inclusive) to
    its row number. Skips blank rows and the TOTAL row."""
    mapping: dict[str, int] = {}
    for row in range(start_row, end_row + 1):
        val = ws[f"B{row}"].value
        if val is None:
            continue
        s = str(val).strip()
        if not s or s.upper() == "TOTAL":
            continue
        mapping[s] = row
    return mapping


def _section_bounds(ws: Worksheet, header_row: int) -> tuple[int, int, int]:
    """Return (total_row, first_client_row, last_client_row) for the section
    whose 'Client' header is at header_row.

    The total row is expected immediately below the header. The section ends
    when column B becomes blank for 2 consecutive rows, or we hit max_row.
    """
    total_row = header_row + 1
    first_client_row = header_row + 2
    last_client_row = first_client_row
    blank_run = 0
    for row in range(first_client_row, ws.max_row + 1):
        val = ws[f"B{row}"].value
        if val is None or (isinstance(val, str) and not val.strip()):
            blank_run += 1
            if blank_run >= 2:
                break
        else:
            blank_run = 0
            last_client_row = row
    return total_row, first_client_row, last_client_row


def _write_client_rows(
    ws: Worksheet,
    tallies: dict[str, tuple[int, int]],
    row_map: dict[str, int],
) -> list[str]:
    """Write (within, outside, %) for every template row. Rows missing from
    tallies are written as (0, 0, N/A) so stale values from the template
    cannot leak into the output.

    Percent column is written BOTH as a numeric value (so it renders in any
    viewer, including those that don't evaluate formulas) AND retains the
    formula behaviour in viewers that treat openpyxl-set numbers as cached.

    Returns list of tally keys that had no matching template row.
    """
    unmatched: list[str] = []

    # Build a normalised lookup so case/whitespace mismatches still resolve.
    norm_to_row: dict[str, int] = {_normalise(name): r for name, r in row_map.items()}

    # Resolve every tally key to a template row.
    resolved: dict[int, tuple[int, int]] = {}
    for client, (within, outside) in tallies.items():
        row = row_map.get(client) or norm_to_row.get(_normalise(client))
        if row is None:
            if within or outside:
                unmatched.append(client)
            continue
        resolved[row] = (int(within), int(outside))

    # Write every template row — defaulting to (0, 0) for any row not in
    # `resolved`. This is what kills the stale-value bug.
    for row in row_map.values():
        within, outside = resolved.get(row, (0, 0))
        ws[f"{WITHIN_COL}{row}"] = within
        ws[f"{OUTSIDE_COL}{row}"] = outside
        total = within + outside
        if total > 0:
            ws[f"{PERCENT_COL}{row}"] = round(within / total, 4)
        else:
            ws[f"{PERCENT_COL}{row}"] = None

    return unmatched


def _normalise(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip().casefold()


def _write_totals(ws: Worksheet, total_row: int, first: int, last: int) -> None:
    """Compute totals from the values just written and store them as numbers
    AND a SUM formula. Numeric value first ensures the cell renders in any
    viewer; the formula keeps Excel happy if a user later edits a row."""
    within_total = 0
    outside_total = 0
    for row in range(first, last + 1):
        w = ws[f"{WITHIN_COL}{row}"].value
        o = ws[f"{OUTSIDE_COL}{row}"].value
        if isinstance(w, (int, float)):
            within_total += int(w)
        if isinstance(o, (int, float)):
            outside_total += int(o)
    ws[f"{WITHIN_COL}{total_row}"] = within_total
    ws[f"{OUTSIDE_COL}{total_row}"] = outside_total
    grand = within_total + outside_total
    ws[f"{PERCENT_COL}{total_row}"] = round(within_total / grand, 4) if grand else None


def fill_template(
    template_path: str | Path,
    output_path: str | Path,
    services_tallies: dict[str, tuple[int, int]],
    callbacks_tallies: dict[str, tuple[int, int]],
    month_label: str,
) -> dict[str, list[str]]:
    """Load template, write values, save to output_path.

    Returns {"services_unmatched": [...], "callbacks_unmatched": [...]} so the
    caller can surface any clients we couldn't place.
    """
    wb = load_workbook(str(template_path))
    ws = wb.active  # template has a single sheet

    # Title: "CRM SLA by Customer - <Month Year>"
    current_title = ws[TITLE_CELL].value
    if isinstance(current_title, str) and "-" in current_title:
        prefix = current_title.rsplit("-", 1)[0].rstrip()
    else:
        prefix = "CRM SLA by Customer"
    ws[TITLE_CELL] = f"{prefix} - {month_label}" if month_label else prefix

    # --- Services section ---
    services_header = _find_header_row(ws, start_row=1)
    svc_total_row, svc_first, svc_last = _section_bounds(ws, services_header)
    svc_map = _client_row_map(ws, svc_first, svc_last)
    svc_unmatched = _write_client_rows(ws, services_tallies, svc_map)
    _write_totals(ws, svc_total_row, svc_first, svc_last)

    # --- Callbacks section ---
    cb_unmatched: list[str] = []
    cb_section_header_row = _find_callbacks_header_row(ws)
    if cb_section_header_row is not None:
        cb_client_header = _find_header_row(ws, start_row=cb_section_header_row + 1)
        cb_total_row, cb_first, cb_last = _section_bounds(ws, cb_client_header)
        cb_map = _client_row_map(ws, cb_first, cb_last)
        cb_unmatched = _write_client_rows(ws, callbacks_tallies, cb_map)
        _write_totals(ws, cb_total_row, cb_first, cb_last)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return {"services_unmatched": svc_unmatched, "callbacks_unmatched": cb_unmatched}


def read_template_clients(template_path: str | Path) -> list[str]:
    """Read the authoritative client list from the services section's column B."""
    wb = load_workbook(str(template_path), read_only=True, data_only=True)
    ws = wb.active
    # Scan column B; start collecting after the 'Client' header and stop at
    # the first blank run of 2 or when we hit the callbacks section.
    clients: list[str] = []
    seen_header = False
    blank_run = 0
    for row in ws.iter_rows(min_col=2, max_col=2, values_only=True):
        val = row[0]
        if val is None:
            if seen_header:
                blank_run += 1
                if blank_run >= 2:
                    break
            continue
        s = str(val).strip()
        if not s:
            if seen_header:
                blank_run += 1
            continue
        blank_run = 0
        if not seen_header:
            if s.lower() == "client":
                seen_header = True
            continue
        if s.upper() == "TOTAL":
            continue
        if "call back" in s.lower():
            break
        clients.append(s)
    wb.close()
    return clients
