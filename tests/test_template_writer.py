"""Verify the template writer preserves structure and computes percentages
via formulas so Excel recalculates on open.
"""
from pathlib import Path

from openpyxl import load_workbook

import template_writer

ROOT = Path(__file__).parent.parent
TEMPLATE = ROOT / "templates" / "HSBC_SLA_template.xlsx"


def test_read_template_clients_returns_70(tmp_path):
    clients = template_writer.read_template_clients(TEMPLATE)
    assert len(clients) == 70
    assert clients[0] == "**ArdexandBA**"
    assert "Honeywell UK" in clients
    assert "YOUFIBRE LIMITED" in clients


def test_fill_template_writes_values_and_preserves_structure(tmp_path):
    out = tmp_path / "filled.xlsx"
    services = {
        "Honeywell UK": (150, 3),
        "Honeywell International": (290, 8),
        "Ashfield Healthcare": (25, 0),
    }
    callbacks = {"Honeywell International": (10, 0)}

    result = template_writer.fill_template(
        template_path=TEMPLATE,
        output_path=out,
        services_tallies=services,
        callbacks_tallies=callbacks,
        month_label="April 2026",
    )
    assert out.exists()
    assert result["services_unmatched"] == []
    assert result["callbacks_unmatched"] == []

    wb = load_workbook(out, data_only=False)
    ws = wb.active

    # Title updated
    assert "April 2026" in str(ws["B2"].value)

    # Row 7 TOTAL is now written as a numeric value (was a SUM formula
    # before — switched to value because some viewers don't evaluate
    # formulas, which manifested as a blank %-column in production).
    assert ws["C7"].value == 150 + 290 + 25  # within total
    assert ws["D7"].value == 3 + 8 + 0       # outside total

    # Services rows written
    # Find Honeywell UK row
    hw_row = None
    for r in range(8, 78):
        if ws[f"B{r}"].value == "Honeywell UK":
            hw_row = r
            break
    assert hw_row is not None
    assert ws[f"C{hw_row}"].value == 150
    assert ws[f"D{hw_row}"].value == 3
    # Percent column is now a numeric value (0-1), not a formula.
    pct = ws[f"E{hw_row}"].value
    assert isinstance(pct, float)
    assert abs(pct - 150 / 153) < 1e-3

    # Clients not in the tallies should be cleared to (0, 0) so stale
    # template values cannot leak through to the output.
    other_row = None
    for r in range(8, 78):
        if ws[f"B{r}"].value == "4Com":
            other_row = r
            break
    assert other_row is not None
    assert ws[f"B{other_row}"].value == "4Com"
    assert ws[f"C{other_row}"].value == 0
    assert ws[f"D{other_row}"].value == 0

    # Callbacks section: Honeywell International should have 10 in Within
    cb_hw_row = None
    for r in range(86, ws.max_row + 1):
        if ws[f"B{r}"].value == "Honeywell International":
            cb_hw_row = r
            break
    assert cb_hw_row is not None
    assert ws[f"C{cb_hw_row}"].value == 10


def test_fill_template_template_file_unchanged(tmp_path):
    """Template must never be modified in place."""
    original_bytes = TEMPLATE.read_bytes()
    out = tmp_path / "filled.xlsx"
    template_writer.fill_template(
        template_path=TEMPLATE,
        output_path=out,
        services_tallies={"Honeywell UK": (1, 0)},
        callbacks_tallies={},
        month_label="April 2026",
    )
    assert TEMPLATE.read_bytes() == original_bytes


def test_fill_template_unmatched_client_returned(tmp_path):
    out = tmp_path / "filled.xlsx"
    result = template_writer.fill_template(
        template_path=TEMPLATE,
        output_path=out,
        services_tallies={"Not In Template Ltd": (5, 1)},
        callbacks_tallies={},
        month_label="April 2026",
    )
    assert "Not In Template Ltd" in result["services_unmatched"]
